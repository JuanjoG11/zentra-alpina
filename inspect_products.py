import pandas as pd
from openpyxl import load_workbook
import time
from collections import Counter

excel_path = r"C:\Users\Juanjo\Downloads\CUBO_DE_VENTAS_TYM_EJE_AWS_de_2026-05-01_a_2026-05-27_user_60_e3679ba8.xlsx"

print("Opening workbook...")
wb = load_workbook(excel_path, read_only=True, data_only=True)
sheet = wb['CUBO_DE_VENTAS']

headers = []
row_idx = 0
total_rows = 0

grupo_articulo_counter = Counter()
marca_counter = Counter()
familia_counter = Counter()
producto_sample = {}

for row in sheet.iter_rows(values_only=True):
    row_idx += 1
    if row_idx == 1:
        headers = list(row)
        idx_grupo = headers.index('nmGrupoArticulo')
        idx_marca = headers.index('nmTpMarca')
        idx_familia = headers.index('nmTpFamilia')
        idx_producto = headers.index('nmProducto')
        continue
    
    if not any(row):
        continue
        
    total_rows += 1
    grupo = row[idx_grupo]
    marca = row[idx_marca]
    familia = row[idx_familia]
    prod = row[idx_producto]
    
    grupo_articulo_counter[grupo] += 1
    marca_counter[marca] += 1
    familia_counter[familia] += 1
    
    if grupo not in producto_sample:
        producto_sample[grupo] = prod

print("\n--- Unique values in nmGrupoArticulo ---")
for k, v in grupo_articulo_counter.items():
    print(f"  {k}: {v} (Sample: {producto_sample.get(k)})")

print("\n--- Unique values in nmTpMarca (top 20) ---")
for k, v in marca_counter.most_common(20):
    print(f"  {k}: {v}")

print("\n--- Unique values in nmTpFamilia (top 20) ---")
for k, v in familia_counter.most_common(20):
    print(f"  {k}: {v}")
