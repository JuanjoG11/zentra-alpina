from openpyxl import load_workbook
import sys

excel_path = r"C:\Users\Juanjo\Downloads\CUBO_DE_VENTAS_TYM_EJE_AWS_de_2026-06-01_a_2026-06-17_user_60_a46bcd22.xlsx"
print("Loading workbook (headers only)...")
wb = load_workbook(excel_path, read_only=True)
sheet = wb['CUBO_DE_VENTAS']

print("Reading first row as headers:")
for row in sheet.iter_rows(values_only=True):
    print("Full Headers list:")
    print(list(row))
    break
sys.exit(0)
