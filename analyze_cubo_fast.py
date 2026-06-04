from openpyxl import load_workbook
import time
from collections import Counter

excel_path = r"C:\Users\Juanjo\Downloads\CUBO_DE_VENTAS_TYM_EJE_AWS_de_2026-05-01_a_2026-05-27_user_60_e3679ba8.xlsx"

print("Opening workbook in read_only mode...")
t0 = time.time()
wb = load_workbook(excel_path, read_only=True, data_only=True)
sheet = wb['CUBO_DE_VENTAS']
print(f"Workbook opened in {time.time() - t0:.2f} seconds.")

# Read headers
headers = []
row_idx = 0
total_rows = 0

forma_pago_counter = Counter()
afecta_venta_counter = Counter()
estado_planilla_counter = Counter()
motivo_counter = Counter()
proveedor_counter = Counter()
ciudad_counter = Counter()
dates = []

vlrAntesIva_sum = 0.0
vlrTotalconIva_sum = 0.0

print("Streaming rows...")
for row in sheet.iter_rows(values_only=True):
    row_idx += 1
    if row_idx == 1:
        headers = list(row)
        print("Headers found:", headers)
        # get indices for columns
        idx_forma_pago = headers.index('nbFormaPago')
        idx_afecta_venta = headers.index('boAfectaVenta')
        idx_estado_planilla = headers.index('estadoPlanilla')
        idx_motivo = headers.index('motivo')
        idx_proveedor = headers.index('nmProveedor')
        idx_ciudad = headers.index('txCiudad')
        idx_fecha = headers.index('dtFactura')
        idx_antes_iva = headers.index('vlrAntesIva')
        idx_total_iva = headers.index('vlrTotalconIva')
        continue
    
    # Check if row is empty
    if not any(row):
        continue
        
    total_rows += 1
    
    # Gather stats
    forma_pago_counter[row[idx_forma_pago]] += 1
    afecta_venta_counter[row[idx_afecta_venta]] += 1
    estado_planilla_counter[row[idx_estado_planilla]] += 1
    motivo_counter[row[idx_motivo]] += 1
    proveedor_counter[row[idx_proveedor]] += 1
    ciudad_counter[row[idx_ciudad]] += 1
    
    dt = row[idx_fecha]
    if dt:
        dates.append(str(dt))
        
    val_antes = row[idx_antes_iva]
    if val_antes is not None:
        try:
            vlrAntesIva_sum += float(val_antes)
        except:
            pass
            
    val_total = row[idx_total_iva]
    if val_total is not None:
        try:
            vlrTotalconIva_sum += float(val_total)
        except:
            pass

    if total_rows % 50000 == 0:
        print(f"Processed {total_rows} rows...")

print(f"Finished processing {total_rows} rows in {time.time() - t0:.2f} seconds.")

print("\n--- Unique values in nbFormaPago ---")
for k, v in forma_pago_counter.items():
    print(f"  {k}: {v}")

print("\n--- Unique values in boAfectaVenta ---")
for k, v in afecta_venta_counter.items():
    print(f"  {k}: {v}")

print("\n--- Unique values in estadoPlanilla ---")
for k, v in estado_planilla_counter.items():
    print(f"  {k}: {v}")

print("\n--- Motivos de Devolucion (top 15) ---")
for k, v in motivo_counter.most_common(15):
    print(f"  {k}: {v}")

print("\n--- Unique values in nmProveedor ---")
for k, v in proveedor_counter.items():
    print(f"  {k}: {v}")

print("\n--- Date Range ---")
if dates:
    print(f"  Min: {min(dates)}")
    print(f"  Max: {max(dates)}")

print("\n--- Financial Stats ---")
print(f"  Sum of vlrAntesIva: {vlrAntesIva_sum:,.2f}")
print(f"  Sum of vlrTotalconIva: {vlrTotalconIva_sum:,.2f}")

# Write raw summary
summary_path = r"C:\Users\Juanjo\Documents\zentra alpina\data\cubo_summary_fast.txt"
with open(summary_path, 'w', encoding='utf-8') as f:
    f.write(f"Total Rows: {total_rows}\n")
    f.write(f"Date Range: {min(dates) if dates else 'N/A'} to {max(dates) if dates else 'N/A'}\n")
    f.write(f"Forma Pago Counter: {dict(forma_pago_counter)}\n")
    f.write(f"Afecta Venta Counter: {dict(afecta_venta_counter)}\n")
    f.write(f"Estado Planilla Counter: {dict(estado_planilla_counter)}\n")
    f.write(f"Motivo Counter: {dict(motivo_counter.most_common(50))}\n")
    f.write(f"Proveedor Counter: {dict(proveedor_counter)}\n")
    f.write(f"Ciudad Counter: {dict(ciudad_counter)}\n")
    f.write(f"Sum of vlrAntesIva: {vlrAntesIva_sum}\n")
    f.write(f"Sum of vlrTotalconIva: {vlrTotalconIva_sum}\n")
print(f"Summary written to {summary_path}")
